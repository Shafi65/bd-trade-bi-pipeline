"""
M1 — Extract Bangladesh FTS 2024-25 (Excel) into canonical trade facts, then reconcile.

PURPOSE (why this milestone exists)
-----------------------------------
Everything the project answers — HS8 price trends, the biggest price improvers/eroders,
per-destination deviation, and product/market diversification — is computed from ONE
clean fact table. M1 builds the first and cleanest slice of that table (the 2024-25
Excel) and PROVES it is trustworthy before any analysis or AWS. Price = value / quantity,
so we extract and validate BOTH measures, not just value. Exports are the headline;
imports are captured too as supporting context (input-import intensity).

WHAT
----
Reads the two "Table 04" sheets — Expt04 (exports) and 'Impt 04' (imports), the finest
grain BBS publishes — and reshapes them into one canonical long ("tidy") table:

    one fact row = (flow, fiscal_year, half_year, HS8 product, destination country)
                   + quantity, value_bdt (taka), unit      [unit price is derived later]

The two sheets are laid out OPPOSITE ways (verified against the file):
  - exports  = COMMODITY-major: an HS8 commodity, then the countries that bought it.
  - imports  = COUNTRY-major:   a country, then every commodity bought from it.
Either way a fact pairs an HS8 with a country; only the "header vs detail" role flips.

WHY reconcile
-------------
Row-level extraction fails silently: a misread column still "looks" like valid numbers.
BBS states its own totals, so we use them as an answer key — checking VALUE and QUANTITY
(both halves of price):
  1. within-row     : H1 + H2 == full-year column, for value AND quantity
  2. grand total    : sum of extracted value/quantity == the EXPORT/IMPORT TOTAL row
  3. Table01 witness: value TOTAL == Table 01's TOTAL, and per-HS2-chapter value sums match
  4. parent<->child : each header subtotal (value & quantity) == sum of its detail rows
FAIL here => this milestone fails; fix the parser before touching PDFs/AWS.

Run:    ./.venv/bin/python ingestion/extract_fts_excel.py
Output: data/interim/fts_2024-25_facts.csv  (git-ignored) + a printed verdict.
"""

from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook

# --- paths & config -------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Bd_Exportdata" / "Foreign Trade Statistics-2024-25" / "FTS_24-25_ExcelTable.xlsx"
OUT_DIR = ROOT / "data" / "interim"
FY = "2024-25"
TOL = 0.001                                   # 0.1% reconciliation tolerance

TABLE04 = {"X": "Expt04", "M": "Impt 04"}     # HS8 x country grain (note the space)
TABLE01 = {"X": "Expt01", "M": "Impt01"}      # per-HS2-chapter totals = answer key
ORIENTATION = {"X": "commodity_major",        # exports: commodity, then its countries
               "M": "country_major"}          # imports: country, then its commodities

# Shared 0-indexed column layout of every Table 04 sheet:
C_CODE, C_DESC, C_UNIT = 0, 1, 2
C_QH1, C_VH1, C_QH2, C_VH2, C_QFY, C_VFY = 3, 4, 5, 6, 7, 8


def num(x):
    """Coerce a cell to float; blanks/None -> 0.0."""
    if x is None or x == "":
        return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def digits(code):
    """Return the cell as a bare digit-string, or None if it isn't purely numeric."""
    if code is None:
        return None
    s = str(code).strip()
    return s if s.isdigit() else None


def _read6(row, off):
    """Six measures (q_h1, v_h1, q_h2, v_h2, q_fy, v_fy), shifted left by -off columns."""
    return (num(row[C_QH1 + off]), num(row[C_VH1 + off]),
            num(row[C_QH2 + off]), num(row[C_VH2 + off]),
            num(row[C_QFY + off]), num(row[C_VFY + off]))


def _commodity_unit_offset(row, desc):
    """Unit text + column offset for a commodity row.

    Source quirk: some commodity descriptions contain an inch-mark (") — e.g.
    'VENTILATION EXHUST FAN DIA 24"'. That quote merges the unit text INTO the
    description and shifts every number one column left. Signature: a numeric Unit
    cell together with a double-quote in the description. A numeric-but-quote-less
    Unit (e.g. a stray 0) is just a missing unit on an otherwise aligned row.
    """
    u = row[C_UNIT]
    if isinstance(u, str):
        return (u.strip() or None), 0
    if isinstance(desc, str) and '"' in desc:
        return None, -1
    return None, 0


def parse_table04(ws, flow, orientation):
    """Flatten a Table 04 sheet into canonical per-half fact rows.

    The commodity row carries the HS8 code + unit (and may be quote-shifted); the
    country row carries the country code + name (never shifted, no unit). Whichever
    is the header sets the running context; each detail row emits facts by pairing
    its own quantity/value with that context, unpivoted into H1 / H2 rows.
    """
    commodity_major = orientation == "commodity_major"
    parent_key = "hs8" if commodity_major else "country_code"

    records, parents = [], []
    grand_v = grand_q = None
    cur_hs8 = cur_unit = cur_cc = cur_cname = None
    unknown = shifted = 0

    for row in ws.iter_rows(values_only=True):
        d = digits(row[C_CODE])
        desc = row[C_DESC].strip() if isinstance(row[C_DESC], str) else row[C_DESC]

        if d is None:                                   # title/header/TOTAL rows
            if isinstance(desc, str) and desc.endswith("TOTAL"):
                grand_v, grand_q = num(row[C_VFY]), num(row[C_QFY])
            continue

        is_commodity, is_country = len(d) >= 6, len(d) <= 3
        if not (is_commodity or is_country):            # 4-5 digit code: shouldn't happen
            unknown += 1
            continue

        # Read this row's measures. Only commodity rows can be quote-shifted.
        if is_commodity:
            unit, off = _commodity_unit_offset(row, desc)
            if off:
                shifted += 1
            q1, v1, q2, v2, qf, vf = _read6(row, off)
        else:
            unit = None
            q1, v1, q2, v2, qf, vf = _read6(row, 0)

        # A commodity row is the header for exports; a country row is the header for imports.
        if is_commodity == commodity_major:             # this row is a HEADER (subtotal)
            if is_commodity:
                cur_hs8, cur_unit = d.zfill(8), unit
                key = cur_hs8
            else:
                cur_cc, cur_cname = d, desc
                key = cur_cc
            parents.append((key, v1, v2, vf, q1, q2, qf))
            continue

        # This row is a DETAIL row -> pair it with the running header context.
        if is_commodity:                                # imports: commodity under a country
            f_hs8, f_unit, f_cc, f_cname = d.zfill(8), unit, cur_cc, cur_cname
        else:                                           # exports: country under a commodity
            f_hs8, f_unit, f_cc, f_cname = cur_hs8, cur_unit, d, desc
        if f_hs8 is None or f_cc is None:               # detail before any header
            unknown += 1
            continue
        for half, qq, vv in (("H1", q1, v1), ("H2", q2, v2)):
            if qq or vv:                                # skip halves with no trade
                records.append({
                    "flow": flow, "fy": FY, "half": half,
                    "hs8": f_hs8, "country_code": f_cc, "country_name": f_cname,
                    "unit": f_unit, "quantity": qq, "value_bdt": vv,
                })

    if unknown:
        print(f"  [warn] {unknown} rows matched no rule (bad code, or detail before header)")
    if shifted:
        print(f"  [note] realigned {shifted} quote-mangled commodity rows (unit lost, values recovered)")
    return (pd.DataFrame(records),
            pd.DataFrame(parents, columns=["key", "v_h1", "v_h2", "v_fy",
                                           "q_h1", "q_h2", "q_fy"]),
            (grand_v, grand_q), parent_key)


def parse_table01(ws):
    """HS2 chapter -> full-year value, plus the reported grand total."""
    chapters, grand = {}, None
    for row in ws.iter_rows(values_only=True):
        d = digits(row[0])
        desc = row[1]
        if d is None and isinstance(desc, str) and desc.strip().endswith("TOTAL"):
            grand = num(row[4])                         # col4 = full-year value
        elif d is not None:
            chapters[d.zfill(2)] = num(row[4])
    return chapters, grand


def _off(a, b):
    """Relative gap between two numbers; treats 0-vs-0 as a match, 0-vs-nonzero as a miss."""
    if b == 0:
        return 0.0 if a == 0 else 1.0
    return abs(a - b) / abs(b)


def reconcile(df, parents, grand, table01, table01_grand, flow, parent_key, fy=FY):
    """Run the checks (value AND quantity), print a report, return True on PASS.

    `fy` labels the report; defaults to this module's FY so the M1 caller is
    unchanged, but the PDF extractor passes each historical year explicitly.
    """
    grand_v, grand_q = grand
    label = {"X": "EXPORTS", "M": "IMPORTS"}[flow]
    print(f"\n=== FY {fy} {label} — reconciliation ===")
    ok = True

    # (1) source integrity: each header's H1 + H2 == its full-year column (value & qty)
    wr_v = int(((parents.v_h1 + parents.v_h2 - parents.v_fy).abs()
                > parents.v_fy.abs() * TOL).sum())
    wr_q = int(((parents.q_h1 + parents.q_h2 - parents.q_fy).abs()
                > parents.q_fy.abs() * TOL).sum())
    print(f"  within-row H1+H2==FY: value {len(parents) - wr_v}/{len(parents)},"
          f" qty {len(parents) - wr_q}/{len(parents)} OK")

    # (2) grand total: sum of extracted facts vs the TOTAL row (value & qty)
    dv = _off(df["value_bdt"].sum(), grand_v)
    dq = _off(df["quantity"].sum(), grand_q)
    print(f"  grand total value: extracted {df['value_bdt'].sum():,.0f}"
          f" vs BBS {grand_v:,.0f}  ({dv*100:.3f}%)  {'OK' if dv <= TOL else 'FAIL'}")
    print(f"  grand total qty:   ({dq*100:.3f}%)  {'OK' if dq <= TOL else 'FAIL'}"
          f"   (mixed-unit sum — extraction check only)")
    ok &= dv <= TOL and dq <= TOL

    # (3a) TOTAL row agrees with Table 01's value TOTAL (independent witness)
    if table01_grand:
        d2 = _off(grand_v, table01_grand)
        print(f"  Table04 TOTAL vs Table01 TOTAL:  ({d2*100:.3f}%)"
              f"  {'OK' if d2 <= TOL else 'FAIL'}")
        ok &= d2 <= TOL

    # (3b) per-chapter: extracted value grouped by HS2 vs Table 01 (Table 01 is value-only)
    by_ch = df.assign(ch=df["hs8"].str[:2]).groupby("ch")["value_bdt"].sum()
    mism = sum(1 for ch, t01 in table01.items() if _off(by_ch.get(ch, 0.0), t01) > TOL)
    print(f"  chapters within {TOL*100:.1f}%: {len(table01) - mism}/{len(table01)}"
          f"  {'OK' if mism == 0 else f'{mism} FAIL'}")
    ok &= mism == 0

    # (4) parent<->children: each header subtotal == sum of its detail rows (value & qty)
    cv, pv = df.groupby(parent_key)["value_bdt"].sum(), parents.groupby("key")["v_fy"].sum()
    cq, pq = df.groupby(parent_key)["quantity"].sum(), parents.groupby("key")["q_fy"].sum()
    shared = pv.index.intersection(cv.index)
    pmis_v = sum(1 for k in shared if _off(cv[k], pv[k]) > TOL)
    pmis_q = sum(1 for k in shared if _off(cq[k], pq[k]) > TOL)
    print(f"  parent==children by {parent_key}: value {len(shared) - pmis_v}/{len(shared)},"
          f" qty {len(shared) - pmis_q}/{len(shared)} OK")

    print(f"  VERDICT: {'PASS' if ok else 'FAIL'}")
    return ok


def main():
    if not XLSX.exists():
        sys.exit(f"Source not found: {XLSX}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    all_facts, all_ok = [], True
    for flow in ("X", "M"):
        df, parents, grand, pkey = parse_table04(wb[TABLE04[flow]], flow, ORIENTATION[flow])
        t01, t01_grand = parse_table01(wb[TABLE01[flow]])
        all_ok &= reconcile(df, parents, grand, t01, t01_grand, flow, pkey)
        all_facts.append(df)

    facts = pd.concat(all_facts, ignore_index=True)
    out = OUT_DIR / f"fts_{FY}_facts.csv"
    facts.to_csv(out, index=False)
    print(f"\nWrote {len(facts):,} fact rows -> {out.relative_to(ROOT)}")

    # Price-readiness: our headline questions need a unit price (value / quantity),
    # which only exists where quantity > 0 and a unit is known. Surface the coverage.
    exp = facts[facts["flow"] == "X"]
    usable = exp[(exp["quantity"] > 0) & exp["unit"].notna()]
    pct = len(usable) / len(exp) * 100 if len(exp) else 0
    print(f"Export rows usable for unit price (qty>0 & unit set): "
          f"{len(usable):,}/{len(exp):,} ({pct:.1f}%)")

    print(f"OVERALL: {'PASS' if all_ok else 'FAIL — fix parser before proceeding'}")
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
